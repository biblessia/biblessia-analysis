#!/usr/bin/env python3
"""Excel 데이터를 읽어 HTML 리포트 생성"""

import json
import sys
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook

# 프로젝트 루트 경로
PROJECT_ROOT = Path(__file__).parent.parent
REPORTS_DIR = PROJECT_ROOT / "reports"


def get_week_title():
    """현재 월 기준으로 'X월 보고서' 타이틀 생성"""
    month = datetime.now().month
    return f"{month}월 보고서"


def find_latest_excel():
    """최신 Excel 파일 찾기"""
    excel_files = sorted(REPORTS_DIR.glob("amplitude_report_*.xlsx"), reverse=True)
    if not excel_files:
        raise FileNotFoundError("No Excel files found in reports/")
    return excel_files[0]


def extract_timeseries(ws, start_row=4, exclude_last=True):
    """시계열 데이터 추출 (WAU, NAU)

    Args:
        exclude_last: True면 마지막 데이터 제외 (이번 주는 수집 중이므로)
    """
    data = {"dates": [], "values": []}
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if row[0] and row[1] is not None:
            data["dates"].append(str(row[0]))
            data["values"].append(row[1])

    # 이번 주 데이터 제외 (수집 중)
    if exclude_last and data["dates"]:
        data["excluded_date"] = data["dates"].pop()
        data["excluded_value"] = data["values"].pop()

    return data


def extract_retention(ws, exclude_last=True):
    """리텐션 데이터 추출

    Args:
        exclude_last: True면 최신 코호트 제외 (이번 주는 수집 중이므로)
    """
    data = {"headers": [], "rows": [], "excluded_cohort": None}
    for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        if row[0]:
            row_data = [cell if cell is not None else "" for cell in row]
            if idx == 0:
                data["headers"] = row_data
            else:
                data["rows"].append(row_data)

    # 이번 주 코호트 제외 (Overall 행 제외하고 첫 번째 코호트가 최신)
    if exclude_last and len(data["rows"]) > 2:
        # Overall Retained, Overall Retained % 다음이 최신 코호트
        for i, row in enumerate(data["rows"]):
            if "Overall" not in str(row[1]):
                data["excluded_cohort"] = data["rows"].pop(i)
                break

    return data


def extract_all_data(excel_path):
    """모든 시트에서 데이터 추출"""
    wb = load_workbook(excel_path, data_only=True)

    data = {
        "file": excel_path.name,
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "wau": None,
        "nau": None,
        "retention": None
    }

    if "WAU" in wb.sheetnames:
        data["wau"] = extract_timeseries(wb["WAU"])

    if "NAU" in wb.sheetnames:
        data["nau"] = extract_timeseries(wb["NAU"])

    if "Weekly Retention" in wb.sheetnames:
        data["retention"] = extract_retention(wb["Weekly Retention"])

    return data


def generate_html(data, insights=None, title=None):
    """HTML 리포트 생성 (Dark Theme)

    Args:
        data: 추출된 데이터
        insights: 인사이트 딕셔너리 (없으면 placeholder)
        title: 보고서 타이틀 (없으면 자동 생성)
    """
    # 타이틀 자동 생성
    if title is None:
        title = get_week_title()

    # 기본 인사이트 (placeholder)
    if insights is None:
        insights = {
            "summary": "<!-- SUMMARY_INSIGHT -->",
            "wau": "<!-- WAU_INSIGHT -->",
            "nau": "<!-- NAU_INSIGHT -->",
            "retention": "<!-- RETENTION_INSIGHT -->",
            "retention_over_time": "<!-- RETENTION_OVER_TIME_INSIGHT -->"
        }

    # WAU/NAU 차트 데이터 - 날짜 포맷 간소화
    wau_labels_short = []
    nau_labels_short = []
    if data["wau"]:
        for d in data["wau"]["dates"]:
            # "2025-07-28" -> "Jul 28"
            try:
                dt = datetime.strptime(d, "%Y-%m-%d")
                wau_labels_short.append(dt.strftime("%b %d"))
            except:
                wau_labels_short.append(d)
    if data["nau"]:
        for d in data["nau"]["dates"]:
            try:
                dt = datetime.strptime(d, "%Y-%m-%d")
                nau_labels_short.append(dt.strftime("%b %d"))
            except:
                nau_labels_short.append(d)

    wau_labels = json.dumps(wau_labels_short) if data["wau"] else "[]"
    wau_values = json.dumps(data["wau"]["values"]) if data["wau"] else "[]"
    nau_labels = json.dumps(nau_labels_short) if data["nau"] else "[]"
    nau_values = json.dumps(data["nau"]["values"]) if data["nau"] else "[]"

    # Summary 계산
    if data["wau"] and len(data["wau"]["values"]) >= 2:
        latest_wau = data["wau"]["values"][-1]
        prev_wau = data["wau"]["values"][-2]
        wau_change = ((latest_wau - prev_wau) / prev_wau * 100) if prev_wau else 0
    else:
        latest_wau = prev_wau = wau_change = 0

    if data["nau"] and len(data["nau"]["values"]) >= 2:
        latest_nau = data["nau"]["values"][-1]
        prev_nau = data["nau"]["values"][-2]
        nau_change = ((latest_nau - prev_nau) / prev_nau * 100) if prev_nau else 0
    else:
        latest_nau = prev_nau = nau_change = 0

    # Retention Week 1 및 곡선 데이터
    week1_retention = "-"
    week1_retention_val = 0  # 평균 리텐션 값 (비교용)
    latest_cohort_retention = "-"
    latest_cohort_retention_val = 0  # 최근 코호트 리텐션 값 (비교용)
    latest_cohort_diff = 0  # 평균 대비 차이 (pp)
    retention_curve_labels = []
    retention_curve_values = []
    if data["retention"] and len(data["retention"]["rows"]) >= 2:
        # "Overall" "Retained %" 행 찾기
        for row in data["retention"]["rows"]:
            if "Overall" in str(row[1]) and "%" in str(row[2]):
                # 소수점 반올림하여 정수로 표시
                if len(row) > 4 and row[4]:
                    week1_retention_val = float(str(row[4]).replace("%", ""))
                    week1_retention = f"{round(week1_retention_val)}%"
                else:
                    week1_retention = "-"
                # 리텐션 곡선 데이터 추출 (W14까지만)
                for i, val in enumerate(row[3:]):
                    if i > 14:  # W14까지만 (W0 ~ W14 = 15개)
                        break
                    if val and "%" in str(val):
                        retention_curve_labels.append(f"W{i}")
                        # % 제거하고 숫자로 변환
                        retention_curve_values.append(float(str(val).replace("%", "")))
                break

        # 최근 코호트 중 Week 1 데이터가 valid한 것 찾기
        # 가장 최신 코호트의 Week 1은 아직 수집 중이므로 두 번째 코호트를 사용
        valid_week1_cohorts = []
        for row in data["retention"]["rows"]:
            if "Overall" not in str(row[1]) and len(row) > 4 and row[4] and row[4] != "":
                # Week 0 (row[3])과 Week 1 (row[4])이 있는 경우
                if row[3] and row[4] and isinstance(row[3], (int, float)) and isinstance(row[4], (int, float)):
                    valid_week1_cohorts.append(row)

        # 두 번째 코호트 사용 (첫 번째는 아직 수집 중)
        if len(valid_week1_cohorts) >= 2:
            row = valid_week1_cohorts[1]
            latest_cohort_retention_val = round(row[4] / row[3] * 100)
            latest_cohort_retention = f"{latest_cohort_retention_val}%"
            latest_cohort_diff = latest_cohort_retention_val - round(week1_retention_val)

    # Week 1~4 리텐션 추이 데이터 추출 (코호트별)
    # 각 Week별로 valid 데이터가 있는 코호트들의 리텐션 추이
    week_trends = {1: [], 2: [], 3: [], 4: []}
    if data["retention"] and len(data["retention"]["rows"]) >= 2:
        for row in data["retention"]["rows"]:
            # Overall 행 제외
            if "Overall" in str(row[1]):
                continue
            # 최신 코호트 제외 (Week 1 데이터가 없으면 제외)
            if len(row) <= 4 or not row[4]:
                continue

            date = str(row[1]).replace(", 2025", "").replace(", 2026", "")  # "Dec 29, 2025" -> "Dec 29"
            week0 = row[3]  # Week 0 사용자 수

            if not week0 or not isinstance(week0, (int, float)) or week0 == 0:
                continue

            # Week 1~4 리텐션 계산
            for week_num in [1, 2, 3, 4]:
                col_idx = 3 + week_num  # Week 1=4, Week 2=5, Week 3=6, Week 4=7
                if len(row) > col_idx and row[col_idx] and isinstance(row[col_idx], (int, float)):
                    retention_pct = round(row[col_idx] / week0 * 100, 1)
                    week_trends[week_num].append({"date": date, "retention": retention_pct})

    # 각 Week 추이를 오래된 순서로 정렬 (차트 X축)
    # 가장 최신 데이터 포인트는 아직 수집 중이므로 제외
    for week_num in week_trends:
        week_trends[week_num] = list(reversed(week_trends[week_num]))
        if len(week_trends[week_num]) > 0:
            week_trends[week_num] = week_trends[week_num][:-1]  # 마지막(최신) 제외

    # WAU 테이블 행
    wau_table_rows = ""
    if data["wau"]:
        for date, value in zip(data["wau"]["dates"], data["wau"]["values"]):
            wau_table_rows += f"<tr><td>{date}</td><td>{value:,}</td></tr>\n"

    # NAU 테이블 행
    nau_table_rows = ""
    if data["nau"]:
        for date, value in zip(data["nau"]["dates"], data["nau"]["values"]):
            nau_table_rows += f"<tr><td>{date}</td><td>{value:,}</td></tr>\n"

    # Retention 테이블 (W14까지 = 18개 컬럼: Segment, Start Date, Users, W0~W14)
    MAX_RETENTION_COLS = 18  # W14까지
    retention_table = ""
    if data["retention"]:
        # 헤더
        retention_table += "<tr>"
        for h in data["retention"]["headers"][:MAX_RETENTION_COLS]:
            retention_table += f"<th>{h}</th>"
        retention_table += "</tr>\n"

        # 데이터 행
        for row in data["retention"]["rows"]:
            retention_table += "<tr>"
            for i, cell in enumerate(row[:MAX_RETENTION_COLS]):
                if i >= 3 and isinstance(cell, (int, float)):
                    retention_table += f"<td class='retention-cell'>{cell:,}</td>"
                else:
                    retention_table += f"<td>{cell}</td>"
            retention_table += "</tr>\n"

    # 날짜 포맷팅
    # 날짜를 한국어 형식으로 변환 (2026-01-13 -> 2026년 1월 13일 작성)
    date_str = data["generated"].split()[0]
    date_parts = date_str.split("-")
    report_date = f"{date_parts[0]}년 {int(date_parts[1])}월 {int(date_parts[2])}일 작성"

    html = f'''<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title} - {report_date}</title>
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.min.css">
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        :root {{
            --bg-primary: #0a0a0a;
            --bg-secondary: #111111;
            --bg-tertiary: #1a1a1a;
            --bg-card: #141414;
            --border-subtle: #222222;
            --border-light: #333333;
            --text-primary: #ffffff;
            --text-secondary: #a0a0a0;
            --text-muted: #666666;
            --accent-primary: #00d4aa;
            --accent-secondary: #00b894;
            --accent-glow: rgba(0, 212, 170, 0.15);
            --positive: #00d4aa;
            --negative: #ff6b6b;
            --chart-wau: #ffffff;
            --chart-nau: #00d4aa;
            --chart-retention: #a0a0a0;
            --chart-cohort: #00d4aa;
        }}

        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        html {{
            scroll-behavior: smooth;
        }}

        body {{
            font-family: 'Pretendard', -apple-system, BlinkMacSystemFont, sans-serif;
            background: var(--bg-primary);
            color: var(--text-primary);
            line-height: 1.7;
            font-weight: 400;
            letter-spacing: -0.01em;
        }}

        /* Subtle noise texture overlay */
        body::before {{
            content: '';
            position: fixed;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background-image: url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='noise'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23noise)'/%3E%3C/svg%3E");
            opacity: 0.03;
            pointer-events: none;
            z-index: 1000;
        }}

        .container {{
            max-width: 1200px;
            margin: 0 auto;
            padding: 40px 24px;
            position: relative;
        }}

        /* Header */
        header {{
            text-align: center;
            padding: 80px 40px;
            margin-bottom: 60px;
            position: relative;
            border-bottom: 1px solid var(--border-subtle);
        }}

        header::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 50%;
            transform: translateX(-50%);
            width: 1px;
            height: 40px;
            background: linear-gradient(to bottom, transparent, var(--accent-primary));
        }}

        .report-label {{
            font-size: 0.75rem;
            font-weight: 500;
            letter-spacing: 0.2em;
            text-transform: uppercase;
            color: var(--accent-primary);
            margin-bottom: 24px;
            display: inline-block;
        }}

        header h1 {{
            font-size: 3rem;
            font-weight: 700;
            letter-spacing: -0.03em;
            margin-bottom: 16px;
            background: linear-gradient(135deg, var(--text-primary) 0%, var(--text-secondary) 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }}

        header .meta {{
            font-size: 0.875rem;
            color: var(--text-muted);
            font-weight: 400;
        }}

        header .meta span {{
            margin: 0 12px;
            opacity: 0.5;
        }}

        /* Sections */
        .section {{
            background: var(--bg-card);
            border: 1px solid var(--border-subtle);
            border-radius: 16px;
            padding: 40px;
            margin-bottom: 32px;
            position: relative;
            overflow: hidden;
        }}

        .section::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 1px;
            background: linear-gradient(90deg, transparent, var(--border-light), transparent);
        }}

        .section h2 {{
            font-size: 1.5rem;
            font-weight: 600;
            letter-spacing: -0.02em;
            color: var(--text-primary);
            margin-bottom: 32px;
            padding-bottom: 16px;
            border-bottom: 1px solid var(--border-subtle);
            display: flex;
            align-items: center;
            gap: 12px;
        }}

        .section h2::before {{
            content: '';
            width: 4px;
            height: 20px;
            background: var(--accent-primary);
            border-radius: 2px;
        }}

        /* Summary Grid */
        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 24px;
            margin-bottom: 40px;
        }}

        @media (max-width: 768px) {{
            .summary-grid {{
                grid-template-columns: 1fr;
            }}
        }}

        .metric-card {{
            background: var(--bg-tertiary);
            border: 1px solid var(--border-subtle);
            border-radius: 12px;
            padding: 28px;
            text-align: center;
            position: relative;
            transition: all 0.3s ease;
        }}

        .metric-card:hover {{
            border-color: var(--border-light);
            transform: translateY(-2px);
        }}

        .metric-card .value {{
            font-size: 2.75rem;
            font-weight: 700;
            letter-spacing: -0.03em;
            color: var(--text-primary);
            line-height: 1.2;
            margin-bottom: 8px;
        }}

        .metric-card .label {{
            font-size: 0.8rem;
            font-weight: 500;
            letter-spacing: 0.05em;
            text-transform: uppercase;
            color: var(--text-muted);
            margin-bottom: 12px;
        }}

        .metric-card .change {{
            font-size: 0.875rem;
            font-weight: 500;
            padding: 4px 12px;
            border-radius: 20px;
            display: inline-block;
        }}

        .change.positive {{
            color: var(--positive);
            background: rgba(0, 212, 170, 0.1);
        }}

        .change.negative {{
            color: var(--negative);
            background: rgba(255, 107, 107, 0.1);
        }}

        .change.neutral {{
            color: var(--text-secondary);
            background: var(--bg-secondary);
        }}

        /* Insight Box */
        .insight-box {{
            background: var(--bg-tertiary);
            border: 1px solid var(--accent-primary);
            border-radius: 12px;
            padding: 28px;
            margin: 32px 0;
            position: relative;
            box-shadow: 0 0 40px var(--accent-glow);
        }}

        .insight-box::before {{
            content: '';
            position: absolute;
            top: -1px;
            left: 20%;
            right: 20%;
            height: 1px;
            background: linear-gradient(90deg, transparent, var(--accent-primary), transparent);
        }}

        .insight-box h3 {{
            font-size: 0.75rem;
            font-weight: 600;
            letter-spacing: 0.15em;
            text-transform: uppercase;
            color: var(--accent-primary);
            margin-bottom: 20px;
            display: flex;
            align-items: center;
            gap: 8px;
        }}

        .insight-box h3::before {{
            content: '';
            width: 8px;
            height: 8px;
            background: var(--accent-primary);
            border-radius: 50%;
            box-shadow: 0 0 12px var(--accent-primary);
        }}

        .insight-box p {{
            color: var(--text-secondary);
            font-size: 0.95rem;
            line-height: 1.8;
        }}

        .insight-box strong {{
            color: var(--text-primary);
            font-weight: 600;
        }}

        .insight-box ul {{
            margin: 12px 0 12px 20px;
            color: var(--text-secondary);
        }}

        .insight-box li {{
            margin-bottom: 8px;
            line-height: 1.7;
        }}

        .insight-box li strong {{
            color: var(--accent-primary);
        }}

        /* Chart Container */
        .chart-container {{
            position: relative;
            height: 380px;
            margin: 24px 0;
            padding: 20px;
            background: var(--bg-tertiary);
            border: 1px solid var(--border-subtle);
            border-radius: 12px;
        }}

        /* Tables */
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 0.875rem;
        }}

        th, td {{
            padding: 14px 12px;
            text-align: left;
            border-bottom: 1px solid var(--border-subtle);
        }}

        th {{
            font-size: 0.75rem;
            font-weight: 600;
            letter-spacing: 0.05em;
            text-transform: uppercase;
            color: var(--text-muted);
            background: var(--bg-secondary);
        }}

        td {{
            color: var(--text-secondary);
            font-variant-numeric: tabular-nums;
        }}

        tr:hover td {{
            background: var(--bg-tertiary);
            color: var(--text-primary);
        }}

        .retention-cell {{
            text-align: center;
            font-weight: 500;
        }}

        .data-table {{
            max-height: 400px;
            overflow-y: auto;
            border: 1px solid var(--border-subtle);
            border-radius: 8px;
        }}

        .data-table::-webkit-scrollbar {{
            width: 6px;
        }}

        .data-table::-webkit-scrollbar-track {{
            background: var(--bg-secondary);
        }}

        .data-table::-webkit-scrollbar-thumb {{
            background: var(--border-light);
            border-radius: 3px;
        }}

        /* Collapsible */
        .collapsible {{
            cursor: pointer;
            padding: 14px 20px;
            background: var(--bg-tertiary);
            border: 1px solid var(--border-subtle);
            border-radius: 8px;
            width: 100%;
            text-align: left;
            font-family: 'Pretendard', sans-serif;
            font-size: 0.875rem;
            font-weight: 500;
            color: var(--text-secondary);
            margin-top: 24px;
            transition: all 0.2s ease;
            display: flex;
            align-items: center;
            gap: 10px;
        }}

        .collapsible::before {{
            content: '+';
            font-size: 1.1rem;
            color: var(--accent-primary);
            font-weight: 300;
            transition: transform 0.2s ease;
        }}

        .collapsible:hover {{
            background: var(--bg-secondary);
            border-color: var(--border-light);
            color: var(--text-primary);
        }}

        .collapsible-content {{
            display: none;
            padding-top: 20px;
            animation: fadeIn 0.3s ease;
        }}

        .collapsible-content.active {{
            display: block;
        }}

        @keyframes fadeIn {{
            from {{ opacity: 0; transform: translateY(-10px); }}
            to {{ opacity: 1; transform: translateY(0); }}
        }}

        /* Footer */
        footer {{
            text-align: center;
            padding: 60px 40px;
            border-top: 1px solid var(--border-subtle);
            margin-top: 40px;
        }}

        footer p {{
            font-size: 0.8rem;
            color: var(--text-muted);
            letter-spacing: 0.02em;
        }}

        footer .logo {{
            font-weight: 700;
            color: var(--accent-primary);
        }}

        /* Animations */
        .section {{
            animation: slideUp 0.6s ease forwards;
            opacity: 0;
        }}

        .section:nth-child(1) {{ animation-delay: 0.1s; }}
        .section:nth-child(2) {{ animation-delay: 0.2s; }}
        .section:nth-child(3) {{ animation-delay: 0.3s; }}
        .section:nth-child(4) {{ animation-delay: 0.4s; }}
        .section:nth-child(5) {{ animation-delay: 0.5s; }}

        @keyframes slideUp {{
            from {{
                opacity: 0;
                transform: translateY(30px);
            }}
            to {{
                opacity: 1;
                transform: translateY(0);
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <div class="report-label">비블레시아 월간 보고서</div>
            <h1>{title}</h1>
            <p class="meta">{report_date}</p>
        </header>

        <!-- 핵심 요약 -->
        <div class="section">
            <h2>핵심 요약</h2>
            <div class="summary-grid">
                <div class="metric-card">
                    <div class="value">{latest_wau:,}명</div>
                    <div class="label">최신 WAU</div>
                    <div class="change {"positive" if wau_change >= 0 else "negative"}">전주 대비 {wau_change:+.1f}%</div>
                </div>
                <div class="metric-card">
                    <div class="value">{latest_nau:,}명</div>
                    <div class="label">최신 NAU</div>
                    <div class="change {"positive" if nau_change >= 0 else "negative"}">전주 대비 {nau_change:+.1f}%</div>
                </div>
                <div class="metric-card">
                    <div class="value">{week1_retention}</div>
                    <div class="label">1주차 리텐션 평균</div>
                    <div class="change neutral">16주 평균</div>
                </div>
                <div class="metric-card">
                    <div class="value">{latest_cohort_retention}</div>
                    <div class="label">최근 1주차 리텐션</div>
                    <div class="change {"positive" if latest_cohort_diff >= 0 else "negative"}">평균 대비 {latest_cohort_diff:+d}%p</div>
                </div>
            </div>
            <div class="insight-box">
                <h3>핵심 인사이트</h3>
                <div id="summary-insight">{insights["summary"]}</div>
            </div>
        </div>

        <!-- WAU 섹션 -->
        <div class="section">
            <h2>WAU (주간 활성 사용자)</h2>
            <div class="chart-container">
                <canvas id="wauChart"></canvas>
            </div>
            <div class="insight-box">
                <h3>WAU 분석</h3>
                <div id="wau-insight">{insights["wau"]}</div>
            </div>
            <button class="collapsible" onclick="toggleCollapsible(this)">원본 데이터 보기</button>
            <div class="collapsible-content">
                <div class="data-table">
                    <table>
                        <thead><tr><th>날짜</th><th>값</th></tr></thead>
                        <tbody>{wau_table_rows}</tbody>
                    </table>
                </div>
            </div>
        </div>

        <!-- NAU 섹션 -->
        <div class="section">
            <h2>NAU (주간 신규 사용자)</h2>
            <div class="chart-container">
                <canvas id="nauChart"></canvas>
            </div>
            <div class="insight-box">
                <h3>NAU 분석</h3>
                <div id="nau-insight">{insights["nau"]}</div>
            </div>
            <button class="collapsible" onclick="toggleCollapsible(this)">원본 데이터 보기</button>
            <div class="collapsible-content">
                <div class="data-table">
                    <table>
                        <thead><tr><th>날짜</th><th>값</th></tr></thead>
                        <tbody>{nau_table_rows}</tbody>
                    </table>
                </div>
            </div>
        </div>

        <!-- 주간 리텐션 섹션 -->
        <div class="section">
            <h2>주간 리텐션</h2>
            <div class="chart-container">
                <canvas id="retentionCurveChart"></canvas>
            </div>
            <div class="insight-box">
                <h3>리텐션 분석</h3>
                <div id="retention-insight">{insights["retention"]}</div>
            </div>
            <button class="collapsible" onclick="toggleCollapsible(this)">원본 데이터 보기</button>
            <div class="collapsible-content">
                <div class="data-table">
                    <table>
                        {retention_table}
                    </table>
                </div>
            </div>
        </div>

        <!-- 코호트별 리텐션 추이 섹션 -->
        <div class="section">
            <h2>코호트별 리텐션 추이</h2>
            <div class="chart-container">
                <canvas id="retentionChart"></canvas>
            </div>
            <div class="insight-box">
                <h3>리텐션 트렌드 분석</h3>
                <div id="retention-over-time-insight">{insights["retention_over_time"]}</div>
            </div>
        </div>

        <footer>
            <p><span class="logo">Biblessia Analytics</span> 제작</p>
        </footer>
    </div>

    <script>
        // 접기/펼치기 토글
        function toggleCollapsible(btn) {{
            const content = btn.nextElementSibling;
            content.classList.toggle('active');
        }}

        // Chart.js 다크 테마 설정
        Chart.defaults.color = '#666666';
        Chart.defaults.borderColor = '#222222';
        Chart.defaults.font.family = "'Pretendard', sans-serif";

        // WAU 차트
        new Chart(document.getElementById('wauChart'), {{
            type: 'line',
            data: {{
                labels: {wau_labels},
                datasets: [{{
                    label: 'WAU',
                    data: {wau_values},
                    borderColor: '#ffffff',
                    backgroundColor: 'rgba(255, 255, 255, 0.05)',
                    fill: true,
                    tension: 0.4,
                    borderWidth: 2,
                    pointBackgroundColor: '#ffffff',
                    pointBorderColor: '#0a0a0a',
                    pointBorderWidth: 2,
                    pointRadius: 0,
                    pointHoverRadius: 6
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                interaction: {{
                    intersect: false,
                    mode: 'index'
                }},
                plugins: {{
                    legend: {{ display: false }},
                    tooltip: {{
                        backgroundColor: '#1a1a1a',
                        titleColor: '#ffffff',
                        bodyColor: '#a0a0a0',
                        borderColor: '#333333',
                        borderWidth: 1,
                        cornerRadius: 8,
                        padding: 12
                    }}
                }},
                scales: {{
                    x: {{
                        grid: {{ color: '#1a1a1a' }},
                        ticks: {{ maxRotation: 45, font: {{ size: 11 }} }}
                    }},
                    y: {{
                        beginAtZero: true,
                        grid: {{ color: '#1a1a1a' }},
                        ticks: {{ font: {{ size: 11 }} }}
                    }}
                }}
            }}
        }});

        // NAU 차트
        new Chart(document.getElementById('nauChart'), {{
            type: 'line',
            data: {{
                labels: {nau_labels},
                datasets: [{{
                    label: 'NAU',
                    data: {nau_values},
                    borderColor: '#00d4aa',
                    backgroundColor: 'rgba(0, 212, 170, 0.08)',
                    fill: true,
                    tension: 0.4,
                    borderWidth: 2,
                    pointBackgroundColor: '#00d4aa',
                    pointBorderColor: '#0a0a0a',
                    pointBorderWidth: 2,
                    pointRadius: 0,
                    pointHoverRadius: 6
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                interaction: {{
                    intersect: false,
                    mode: 'index'
                }},
                plugins: {{
                    legend: {{ display: false }},
                    tooltip: {{
                        backgroundColor: '#1a1a1a',
                        titleColor: '#ffffff',
                        bodyColor: '#a0a0a0',
                        borderColor: '#333333',
                        borderWidth: 1,
                        cornerRadius: 8,
                        padding: 12
                    }}
                }},
                scales: {{
                    x: {{
                        grid: {{ color: '#1a1a1a' }},
                        ticks: {{ maxRotation: 45, font: {{ size: 11 }} }}
                    }},
                    y: {{
                        beginAtZero: true,
                        grid: {{ color: '#1a1a1a' }},
                        ticks: {{ font: {{ size: 11 }} }}
                    }}
                }}
            }}
        }});

        // 주간 리텐션 곡선 (전체 주차별 리텐션)
        new Chart(document.getElementById('retentionCurveChart'), {{
            type: 'line',
            data: {{
                labels: {json.dumps(retention_curve_labels)},
                datasets: [{{
                    label: '리텐션 %',
                    data: {json.dumps(retention_curve_values)},
                    borderColor: '#a0a0a0',
                    backgroundColor: 'rgba(160, 160, 160, 0.05)',
                    fill: true,
                    tension: 0.4,
                    borderWidth: 2,
                    pointBackgroundColor: '#a0a0a0',
                    pointBorderColor: '#0a0a0a',
                    pointBorderWidth: 2,
                    pointRadius: 4,
                    pointHoverRadius: 7
                }}]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                interaction: {{
                    intersect: false,
                    mode: 'index'
                }},
                plugins: {{
                    legend: {{ display: false }},
                    tooltip: {{
                        backgroundColor: '#1a1a1a',
                        titleColor: '#ffffff',
                        bodyColor: '#a0a0a0',
                        borderColor: '#333333',
                        borderWidth: 1,
                        cornerRadius: 8,
                        padding: 12,
                        callbacks: {{
                            label: function(context) {{
                                return context.parsed.y + '%';
                            }}
                        }}
                    }}
                }},
                scales: {{
                    x: {{
                        grid: {{ color: '#1a1a1a' }}
                    }},
                    y: {{
                        beginAtZero: true,
                        max: 100,
                        grid: {{ color: '#1a1a1a' }},
                        ticks: {{
                            callback: function(value) {{
                                return value + '%';
                            }}
                        }}
                    }}
                }}
            }}
        }});

        // 코호트별 리텐션 추이 차트 (Week 1~4 멀티라인)
        const weekTrends = {json.dumps(week_trends)};

        // X축 레이블은 Week 1 기준 (가장 많은 데이터)
        const allLabels = weekTrends[1].map(d => d.date);

        // 각 Week 데이터를 레이블에 맞춰 정렬 (없는 데이터는 null)
        function alignData(weekData, labels) {{
            const dataMap = new Map(weekData.map(d => [d.date, d.retention]));
            return labels.map(label => dataMap.get(label) ?? null);
        }}

        new Chart(document.getElementById('retentionChart'), {{
            type: 'line',
            data: {{
                labels: allLabels,
                datasets: [
                    {{
                        label: 'Week 1',
                        data: alignData(weekTrends[1], allLabels),
                        borderColor: '#00d4aa',
                        backgroundColor: 'rgba(0, 212, 170, 0.1)',
                        borderWidth: 2,
                        tension: 0.3,
                        pointRadius: 4,
                        pointHoverRadius: 6,
                        fill: false
                    }},
                    {{
                        label: 'Week 2',
                        data: alignData(weekTrends[2], allLabels),
                        borderColor: '#ffd700',
                        backgroundColor: 'rgba(255, 215, 0, 0.1)',
                        borderWidth: 2,
                        tension: 0.3,
                        pointRadius: 4,
                        pointHoverRadius: 6,
                        fill: false
                    }},
                    {{
                        label: 'Week 3',
                        data: alignData(weekTrends[3], allLabels),
                        borderColor: '#ff6b6b',
                        backgroundColor: 'rgba(255, 107, 107, 0.1)',
                        borderWidth: 2,
                        tension: 0.3,
                        pointRadius: 4,
                        pointHoverRadius: 6,
                        fill: false
                    }},
                    {{
                        label: 'Week 4',
                        data: alignData(weekTrends[4], allLabels),
                        borderColor: '#4ecdc4',
                        backgroundColor: 'rgba(78, 205, 196, 0.1)',
                        borderWidth: 2,
                        tension: 0.3,
                        pointRadius: 4,
                        pointHoverRadius: 6,
                        fill: false
                    }}
                ]
            }},
            options: {{
                responsive: true,
                maintainAspectRatio: false,
                spanGaps: false,
                plugins: {{
                    legend: {{
                        display: true,
                        position: 'top',
                        labels: {{
                            color: '#a0a0a0',
                            usePointStyle: true,
                            pointStyle: 'circle',
                            padding: 20
                        }}
                    }},
                    tooltip: {{
                        backgroundColor: '#1a1a1a',
                        titleColor: '#ffffff',
                        bodyColor: '#a0a0a0',
                        borderColor: '#333333',
                        borderWidth: 1,
                        cornerRadius: 8,
                        padding: 12,
                        callbacks: {{
                            label: function(context) {{
                                if (context.parsed.y === null) return null;
                                return context.dataset.label + ': ' + context.parsed.y + '%';
                            }}
                        }}
                    }}
                }},
                scales: {{
                    x: {{
                        grid: {{ display: false }},
                        ticks: {{ maxRotation: 45, font: {{ size: 10 }}, color: '#a0a0a0' }}
                    }},
                    y: {{
                        beginAtZero: true,
                        max: 100,
                        grid: {{ color: '#1a1a1a' }},
                        ticks: {{
                            color: '#a0a0a0',
                            callback: function(value) {{
                                return value + '%';
                            }}
                        }}
                    }}
                }}
            }}
        }});
    </script>
</body>
</html>
'''
    return html


def main():
    # 최신 Excel 파일 찾기
    excel_path = find_latest_excel()
    print(f"Reading: {excel_path}")

    # 데이터 추출
    data = extract_all_data(excel_path)

    # 커맨드라인 옵션 파싱
    title = None
    json_mode = False

    i = 1
    while i < len(sys.argv):
        if sys.argv[i] == "-j":
            json_mode = True
        elif sys.argv[i] == "--title" and i + 1 < len(sys.argv):
            title = sys.argv[i + 1]
            i += 1
        i += 1

    # JSON 모드
    if json_mode:
        print(json.dumps(data, indent=2, ensure_ascii=False))
        return

    # 타이틀 자동 생성 (지정되지 않은 경우)
    if title is None:
        title = get_week_title()
    print(f"Report title: {title}")

    # HTML 생성
    html = generate_html(data, title=title)

    # 파일 저장
    today = datetime.now().strftime("%Y-%m-%d")
    output_path = REPORTS_DIR / f"analysis_report_{today}.html"
    output_path.write_text(html, encoding="utf-8")
    print(f"Report saved: {output_path}")

    return data


if __name__ == "__main__":
    main()
