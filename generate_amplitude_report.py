#!/usr/bin/env python3
"""Amplitude 데이터를 Excel로 내보내기"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 내보내기 폴더 설정
EXPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")

# 데이터 정의 (Amplitude에서 가져온 데이터 - 2026-07-13 업데이트, 글로벌 기준 / 마지막 미완성주(07-13) 제외)
WAU_DATA = {
    "dates": ["2026-01-26","2026-02-02","2026-02-09","2026-02-16","2026-02-23","2026-03-02","2026-03-09","2026-03-16","2026-03-23","2026-03-30","2026-04-06","2026-04-13","2026-04-20","2026-04-27","2026-05-04","2026-05-11","2026-05-18","2026-05-25","2026-06-01","2026-06-08","2026-06-15","2026-06-22","2026-06-29","2026-07-06"],
    "values": [2193,2344,2297,2220,2470,2804,2983,2945,2997,3046,2992,3143,3035,3034,3089,3099,3057,3082,3139,3568,3602,3603,3753,3914]
}

# WAU 지역별 분석 (한국 vs 한국 외) - 한국: country is South Korea / 한국 외: country is not South Korea
WAU_BY_REGION_DATA = {
    "dates": ["2026-01-26","2026-02-02","2026-02-09","2026-02-16","2026-02-23","2026-03-02","2026-03-09","2026-03-16","2026-03-23","2026-03-30","2026-04-06","2026-04-13","2026-04-20","2026-04-27","2026-05-04","2026-05-11","2026-05-18","2026-05-25","2026-06-01","2026-06-08","2026-06-15","2026-06-22","2026-06-29","2026-07-06"],
    "korea": [2121,2273,2224,2141,2412,2729,2922,2890,2925,2981,2927,3076,2978,2974,2994,2987,2949,2966,3025,3435,3469,3454,3607,3752],
    "non_korea": [110,121,120,117,90,96,76,86,96,89,88,92,84,91,121,138,130,143,134,154,153,194,195,202]
}

NAU_DATA = {
    "dates": ["2026-01-26","2026-02-02","2026-02-09","2026-02-16","2026-02-23","2026-03-02","2026-03-09","2026-03-16","2026-03-23","2026-03-30","2026-04-06","2026-04-13","2026-04-20","2026-04-27","2026-05-04","2026-05-11","2026-05-18","2026-05-25","2026-06-01","2026-06-08","2026-06-15","2026-06-22","2026-06-29","2026-07-06"],
    "values": [341,369,250,245,377,409,508,361,315,329,359,272,236,313,247,291,224,279,279,612,510,434,516,404]
}

RETENTION_DATA = [
    ["Segment", "Start Date", "Users", "Week 0", "Week 1", "Week 2", "Week 3", "Week 4", "Week 5", "Week 6", "Week 7", "Week 8", "Week 9", "Week 10", "Week 11", "Week 12", "Week 13", "Week 14", "Week 15"],
    ["Global", "Overall", "Retained", 5620, 3057, 2233, 1804, 1444, 1071, 902, 788, 662, 519, 432, 343, 258, 197, 103, 63],
    ["Global", "Overall", "Retained %", "100.0%", "58.61%", "47.51%", "42.29%", "38.45%", "34.06%", "31.48%", "30.47%", "28.03%", "25.06%", "23.68%", "22.7%", "20.24%", "19.64%", "15.99%", "20.0%"],
    ["Global", "Jun 29, 2026", 516, 516, 315],
    ["Global", "Jun 22, 2026", 434, 434, 256, 215],
    ["Global", "Jun 15, 2026", 510, 510, 302, 244, 246],
    ["Global", "Jun 08, 2026", 612, 612, 345, 270, 241, 232],
    ["Global", "Jun 01, 2026", 279, 279, 174, 137, 119, 116, 109],
    ["Global", "May 25, 2026", 279, 279, 180, 149, 125, 110, 99, 90],
    ["Global", "May 18, 2026", 224, 224, 132, 110, 102, 99, 88, 80, 79],
    ["Global", "May 11, 2026", 291, 291, 156, 135, 123, 116, 95, 96, 83, 78],
    ["Global", "May 04, 2026", 247, 247, 151, 114, 97, 89, 84, 82, 82, 76, 80],
    ["Global", "Apr 27, 2026", 313, 313, 209, 164, 141, 117, 118, 106, 108, 94, 70, 80],
    ["Global", "Apr 20, 2026", 236, 236, 131, 117, 96, 95, 78, 78, 69, 72, 58, 57, 54],
    ["Global", "Apr 13, 2026", 272, 272, 158, 115, 123, 108, 93, 86, 84, 83, 61, 68, 64, 58],
    ["Global", "Apr 06, 2026", 359, 359, 231, 185, 165, 152, 132, 120, 115, 117, 112, 97, 96, 93, 88],
    ["Global", "Mar 30, 2026", 329, 329, 151, 143, 114, 109, 91, 85, 94, 73, 70, 66, 65, 47, 47, 45],
    ["Global", "Mar 23, 2026", 315, 315, 166, 135, 112, 101, 84, 79, 74, 69, 68, 64, 64, 60, 62, 58, 63],
]

# 스타일 정의
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_workbook():
    wb = Workbook()

    # Summary 시트
    ws_summary = wb.active
    ws_summary.title = "Summary"
    create_summary_sheet(ws_summary)

    # WAU 시트
    ws_wau = wb.create_sheet("WAU")
    create_timeseries_sheet(ws_wau, "Weekly Active Users (WAU)", WAU_DATA)

    # WAU 지역별 시트 (한국 vs 한국 외)
    ws_wau_region = wb.create_sheet("WAU by Region")
    create_region_sheet(ws_wau_region, "WAU by Region (Korea vs Non-Korea)", WAU_BY_REGION_DATA)

    # NAU 시트
    ws_nau = wb.create_sheet("NAU")
    create_timeseries_sheet(ws_nau, "Weekly New Active Users (NAU)", NAU_DATA)

    # Retention 시트
    ws_retention = wb.create_sheet("Weekly Retention")
    create_retention_sheet(ws_retention, RETENTION_DATA)

    return wb

def create_summary_sheet(ws):
    """Summary 시트 생성"""
    ws['A1'] = "Amplitude Report Summary"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')

    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, color="666666")

    # 메트릭 테이블
    headers = ["Metric", "Latest Value", "Previous Week", "Change"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')

    # WAU
    latest_wau = WAU_DATA["values"][-1]
    prev_wau = WAU_DATA["values"][-2]
    wau_change = ((latest_wau - prev_wau) / prev_wau * 100) if prev_wau else 0

    ws.cell(row=5, column=1, value="WAU").border = BORDER
    ws.cell(row=5, column=2, value=latest_wau).border = BORDER
    ws.cell(row=5, column=3, value=prev_wau).border = BORDER
    ws.cell(row=5, column=4, value=f"{wau_change:+.1f}%").border = BORDER

    # NAU
    latest_nau = NAU_DATA["values"][-1]
    prev_nau = NAU_DATA["values"][-2]
    nau_change = ((latest_nau - prev_nau) / prev_nau * 100) if prev_nau else 0

    ws.cell(row=6, column=1, value="NAU").border = BORDER
    ws.cell(row=6, column=2, value=latest_nau).border = BORDER
    ws.cell(row=6, column=3, value=prev_nau).border = BORDER
    ws.cell(row=6, column=4, value=f"{nau_change:+.1f}%").border = BORDER

    # Week 1 Retention
    ws.cell(row=7, column=1, value="Week 1 Retention").border = BORDER
    ws.cell(row=7, column=2, value="58.61%").border = BORDER
    ws.cell(row=7, column=3, value="-").border = BORDER
    ws.cell(row=7, column=4, value="-").border = BORDER

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12

def create_timeseries_sheet(ws, title, data):
    """시계열 데이터 시트 생성"""
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')

    # 헤더
    headers = ["Date", "Value"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')

    # 데이터
    for row, (date, value) in enumerate(zip(data["dates"], data["values"]), 4):
        ws.cell(row=row, column=1, value=date).border = BORDER
        ws.cell(row=row, column=2, value=value).border = BORDER

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12

def create_region_sheet(ws, title, data):
    """지역별 비교 시계열 시트 생성 (한국 vs 한국 외)"""
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    headers = ["Date", "South Korea", "Non-Korea", "Non-Korea %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')

    for row, (date, kr, non_kr) in enumerate(zip(data["dates"], data["korea"], data["non_korea"]), 4):
        total = kr + non_kr
        share = (non_kr / total * 100) if total else 0
        ws.cell(row=row, column=1, value=date).border = BORDER
        ws.cell(row=row, column=2, value=kr).border = BORDER
        ws.cell(row=row, column=3, value=non_kr).border = BORDER
        ws.cell(row=row, column=4, value=f"{share:.2f}%").border = BORDER

    for col, width in [('A', 15), ('B', 14), ('C', 14), ('D', 14)]:
        ws.column_dimensions[col].width = width


def create_retention_sheet(ws, data):
    """리텐션 시트 생성"""
    ws['A1'] = "Weekly Retention (Cohort Analysis)"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    # 도움말 섹션
    help_texts = [
        "[ 읽는 방법 ]",
        "• 각 행은 특정 주에 처음 방문한 사용자 그룹(코호트)입니다.",
        "• Week 0: 첫 방문 주에 활동한 사용자 수 (항상 100%)",
        "• Week 1, 2, 3...: 첫 방문 후 1주, 2주, 3주 뒤에 다시 돌아온 사용자 수",
        "• Overall 행: 전체 코호트의 평균 리텐션율",
        "",
        "예) Dec 22, 2025 코호트가 Week 2에 306명 → 12월 22일 주에 처음 온 465명 중 306명(65.8%)이 2주 후에도 활동"
    ]

    help_start_row = 3
    for i, text in enumerate(help_texts):
        cell = ws.cell(row=help_start_row + i, column=1, value=text)
        if i == 0:
            cell.font = Font(bold=True, color="4472C4")
        else:
            cell.font = Font(color="666666")
        ws.merge_cells(f'A{help_start_row + i}:G{help_start_row + i}')

    # 데이터 작성 (도움말 아래로 이동)
    data_start_row = help_start_row + len(help_texts) + 1
    for row_idx, row_data in enumerate(data, data_start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            if row_idx == data_start_row:  # 헤더
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center')

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    for i in range(4, 21):
        ws.column_dimensions[get_column_letter(i)].width = 10

def main():
    wb = create_workbook()

    # 폴더 생성 (없으면)
    os.makedirs(EXPORT_DIR, exist_ok=True)

    # 파일 저장
    today = datetime.now().strftime('%Y-%m-%d')
    filename = f'amplitude_report_{today}.xlsx'
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    print(f"Excel file created: {filepath}")

if __name__ == "__main__":
    main()
